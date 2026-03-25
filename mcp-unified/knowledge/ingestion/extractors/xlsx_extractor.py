"""
XLSX Extractor

Mengekstrak data dan teks dari Excel spreadsheets.
"""

from pathlib import Path
from typing import Dict, Any, List


class XlsxExtractor:
    """
    Extractor untuk file Excel.
    
    Supports:
        - .xlsx (Excel 2007+)
        - .xls (legacy, via openpyxl/xlrd)
    
    Strategy:
        - Convert setiap sheet ke format text yang readable
        - Preserve struktur tabel
        - Extract metadata sheets
    """
    
    def __init__(self):
        """Initialize XLSX extractor."""
        self._lib = None
    
    def _init_lib(self):
        """Lazy initialization library."""
        if self._lib is None:
            try:
                import openpyxl
                self._lib = "openpyxl"
            except ImportError:
                try:
                    import pandas as pd
                    self._lib = "pandas"
                except ImportError:
                    raise ImportError(
                        "Excel library tidak tersedia. "
                        "Install: pip install openpyxl atau pandas"
                    )
    
    async def extract(self, file_path: str) -> Dict[str, Any]:
        """
        Extract text dan data dari XLSX file.
        
        Args:
            file_path: Path ke .xlsx file
            
        Returns:
            Dict dengan keys:
                - text: Extracted text (formatted as readable text)
                - metadata: Workbook metadata
                - sheets: List of sheet data
        """
        self._init_lib()
        
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File tidak ditemukan: {file_path}")
        
        if self._lib == "openpyxl":
            return await self._extract_with_openpyxl(file_path)
        else:
            return await self._extract_with_pandas(file_path)
    
    async def _extract_with_openpyxl(self, file_path: str) -> Dict[str, Any]:
        """Extract menggunakan openpyxl."""
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, data_only=True, read_only=True)
        
        metadata = {
            "file_type": "xlsx",
            "sheets": wb.sheetnames,
            "sheet_count": len(wb.sheetnames)
        }
        
        text_parts = []
        sheets_data = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
            
            sheet_data = []
            row_count = 0
            
            for row in sheet.iter_rows():
                row_data = []
                row_text_parts = []
                
                for cell in row:
                    value = cell.value
                    if value is not None:
                        cell_text = str(value).strip()
                        if cell_text:
                            row_data.append(cell_text)
                            row_text_parts.append(cell_text)
                    else:
                        row_data.append("")
                
                if row_text_parts:  # Only add non-empty rows
                    sheet_data.append(row_data)
                    text_parts.append(" | ".join(row_text_parts))
                    row_count += 1
                
                # Limit untuk avoid terlalu besar
                if row_count > 1000:
                    text_parts.append("... (truncated, more than 1000 rows)")
                    break
            
            sheets_data.append({
                "name": sheet_name,
                "data": sheet_data,
                "row_count": row_count
            })
        
        wb.close()
        
        full_text = "\n".join(text_parts)
        
        return {
            "text": full_text,
            "metadata": metadata,
            "sheets": sheets_data,
            "ocr_used": False
        }
    
    async def _extract_with_pandas(self, file_path: str) -> Dict[str, Any]:
        """Extract menggunakan pandas."""
        import pandas as pd
        
        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        
        metadata = {
            "file_type": "xlsx",
            "sheets": excel_file.sheet_names,
            "sheet_count": len(excel_file.sheet_names)
        }
        
        text_parts = []
        sheets_data = []
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
            
            # Convert DataFrame ke text
            # Format: header row, then data rows
            sheet_text = df.to_string(index=False)
            text_parts.append(sheet_text)
            
            # Store structured data
            sheets_data.append({
                "name": sheet_name,
                "data": df.fillna("").to_dict(orient='records'),
                "row_count": len(df),
                "column_count": len(df.columns)
            })
        
        full_text = "\n".join(text_parts)
        
        return {
            "text": full_text,
            "metadata": metadata,
            "sheets": sheets_data,
            "ocr_used": False
        }
    
    def _format_sheet_as_text(self, sheet_data: List[List[str]]) -> str:
        """
        Format sheet data ke readable text.
        
        Args:
            sheet_data: List of rows (each row is list of cell values)
            
        Returns:
            Formatted text
        """
        if not sheet_data:
            return ""
        
        # Find max width untuk setiap column
        col_widths = []
        for row in sheet_data:
            for i, cell in enumerate(row):
                if i >= len(col_widths):
                    col_widths.append(0)
                col_widths[i] = max(col_widths[i], len(str(cell)))
        
        # Format rows
        lines = []
        for row in sheet_data:
            formatted_cells = []
            for i, cell in enumerate(row):
                width = col_widths[i] if i < len(col_widths) else 10
                formatted_cells.append(str(cell).ljust(width))
            lines.append(" | ".join(formatted_cells))
        
        return "\n".join(lines)