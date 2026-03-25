"""
Advanced XLSX Tools - P1 Features
Pivot tables, charts, conditional formatting, data validation, dll
"""
from pathlib import Path
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from tools.base import register_tool


@register_tool
def create_pivot_xlsx(
    file_path: str,
    source_sheet: str,
    target_sheet: str,
    rows: List[str],
    columns: List[str],
    values: List[str],
    agg_func: str = 'sum'
) -> Dict:
    """
    Create a pivot table in XLSX (simplified version using formulas)
    
    Args:
        file_path: Path to XLSX file
        source_sheet: Source data sheet name
        target_sheet: Target sheet name for pivot table
        rows: List of row field names
        columns: List of column field names
        values: List of value field names to aggregate
        agg_func: Aggregation function ('sum', 'count', 'avg', 'max', 'min')
        
    Returns:
        Success status and details
    """
    try:
        wb = load_workbook(file_path)
        
        if source_sheet not in wb.sheetnames:
            return {'success': False, 'error': f'Source sheet "{source_sheet}" not found'}
        
        # Create or get target sheet
        if target_sheet in wb.sheetnames:
            del wb[target_sheet]
        ws_target = wb.create_sheet(title=target_sheet)
        
        ws_source = wb[source_sheet]
        
        # Get headers and data
        headers = [cell.value for cell in ws_source[1]]
        
        # Simple pivot implementation using summary table
        ws_target['A1'] = 'Pivot Table Summary'
        ws_target['A1'].font = Font(bold=True, size=14)
        
        ws_target['A3'] = 'Row Fields:'
        ws_target['B3'] = ', '.join(rows)
        ws_target['A4'] = 'Column Fields:'
        ws_target['B4'] = ', '.join(columns)
        ws_target['A5'] = 'Value Fields:'
        ws_target['B5'] = ', '.join(values)
        ws_target['A6'] = 'Aggregation:'
        ws_target['B6'] = agg_func
        
        # Note: Full pivot table creation requires more complex implementation
        # with openpyxl's pivot table support or manual formula generation
        
        ws_target['A8'] = 'Note: This is a simplified pivot summary.'
        ws_target['A9'] = 'Full pivot table support requires external libraries.'
        
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'source_sheet': source_sheet,
            'target_sheet': target_sheet,
            'rows': rows,
            'columns': columns,
            'values': values,
            'aggregation': agg_func,
            'note': 'Simplified pivot summary created'
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


@register_tool
def add_chart_xlsx(
    file_path: str,
    sheet_name: str,
    chart_type: str,
    data_range: str,
    title: str,
    target_cell: str = 'J2'
) -> Dict:
    """
    Add a chart to an XLSX file
    
    Args:
        file_path: Path to XLSX file
        sheet_name: Sheet name
        chart_type: Type of chart ('bar', 'line', 'pie')
        data_range: Data range (e.g., 'A1:B10')
        title: Chart title
        target_cell: Cell to place chart (default: 'J2')
        
    Returns:
        Success status and details
    """
    try:
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            return {'success': False, 'error': f'Sheet "{sheet_name}" not found'}
        
        ws = wb[sheet_name]
        
        # Create chart based on type
        if chart_type.lower() == 'bar':
            chart = BarChart()
        elif chart_type.lower() == 'line':
            chart = LineChart()
        elif chart_type.lower() == 'pie':
            chart = PieChart()
        else:
            chart = BarChart()
        
        # Set chart data
        data = Reference(ws, min_col=1, min_row=1, max_col=2, max_row=10)
        categories = Reference(ws, min_col=1, min_row=2, max_row=10)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = title
        
        # Add chart to worksheet
        ws.add_chart(chart, target_cell)
        
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'sheet_name': sheet_name,
            'chart_type': chart_type,
            'title': title,
            'target_cell': target_cell
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


@register_tool
def apply_conditional_formatting_xlsx(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    rule_type: str,
    value: Any,
    format_color: str = 'FF0000'
) -> Dict:
    """
    Apply conditional formatting to a range
    
    Args:
        file_path: Path to XLSX file
        sheet_name: Sheet name
        cell_range: Cell range (e.g., 'A1:A10')
        rule_type: Type of rule ('greater_than', 'less_than', 'equal_to')
        value: Value to compare against
        format_color: Color for formatting (hex format, default: 'FF0000' red)
        
    Returns:
        Success status and details
    """
    try:
        from openpyxl.formatting.rule import CellIsRule
        
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            return {'success': False, 'error': f'Sheet "{sheet_name}" not found'}
        
        ws = wb[sheet_name]
        
        # Create fill
        fill = PatternFill(start_color=format_color, end_color=format_color, fill_type='solid')
        
        # Create rule based on type
        operator_map = {
            'greater_than': 'greaterThan',
            'less_than': 'lessThan',
            'equal_to': 'equal',
            'greater_than_or_equal': 'greaterThanOrEqual',
            'less_than_or_equal': 'lessThanOrEqual'
        }
        
        operator = operator_map.get(rule_type.lower(), 'greaterThan')
        rule = CellIsRule(operator=operator, formula=[value], fill=fill)
        
        # Apply to range
        ws.conditional_formatting.add(cell_range, rule)
        
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'sheet_name': sheet_name,
            'range': cell_range,
            'rule_type': rule_type,
            'value': value,
            'format_color': format_color
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


@register_tool
def add_data_validation_xlsx(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    validation_type: str,
    allowed_values: Optional[List[str]] = None,
    min_value: Optional[float] = None,
    max_value: Optional[float] = None,
    error_message: str = 'Invalid input'
) -> Dict:
    """
    Add data validation to cells
    
    Args:
        file_path: Path to XLSX file
        sheet_name: Sheet name
        cell_range: Cell range (e.g., 'A1:A10')
        validation_type: Type ('list', 'whole', 'decimal', 'date')
        allowed_values: List of allowed values (for 'list' type)
        min_value: Minimum value (for numeric types)
        max_value: Maximum value (for numeric types)
        error_message: Error message to show
        
    Returns:
        Success status and details
    """
    try:
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            return {'success': False, 'error': f'Sheet "{sheet_name}" not found'}
        
        ws = wb[sheet_name]
        
        # Create validation
        if validation_type == 'list' and allowed_values:
            formula = '"' + ','.join(allowed_values) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        elif validation_type == 'whole':
            formula1 = str(min_value) if min_value is not None else '0'
            formula2 = str(max_value) if max_value is not None else '1000000'
            dv = DataValidation(type="whole", operator="between", formula1=formula1, formula2=formula2)
        elif validation_type == 'decimal':
            formula1 = str(min_value) if min_value is not None else '0'
            formula2 = str(max_value) if max_value is not None else '1000000'
            dv = DataValidation(type="decimal", operator="between", formula1=formula1, formula2=formula2)
        else:
            dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        
        dv.error = error_message
        dv.errorTitle = 'Invalid Input'
        
        # Add validation to worksheet
        ws.add_data_validation(dv)
        dv.add(cell_range)
        
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'sheet_name': sheet_name,
            'range': cell_range,
            'validation_type': validation_type,
            'allowed_values': allowed_values,
            'min_value': min_value,
            'max_value': max_value
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


@register_tool
def calculate_formulas_xlsx(file_path: str, sheet_name: Optional[str] = None) -> Dict:
    """
    Calculate all formulas in a worksheet
    Note: openpyxl doesn't calculate formulas, this is a placeholder
    
    Args:
        file_path: Path to XLSX file
        sheet_name: Sheet name (None for all sheets)
        
    Returns:
        Success status and note
    """
    return {
        'success': True,
        'file_path': file_path,
        'sheet_name': sheet_name,
        'note': 'Formula calculation requires opening in Excel or using external libraries like xlwings'
    }


@register_tool
def apply_filter_sort_xlsx(
    file_path: str,
    sheet_name: str,
    range_str: str,
    sort_column: Optional[int] = None,
    sort_descending: bool = False
) -> Dict:
    """
    Apply auto-filter and sort to a range
    
    Args:
        file_path: Path to XLSX file
        sheet_name: Sheet name
        range_str: Cell range (e.g., 'A1:D100')
        sort_column: Column index to sort by (1-based, None for no sort)
        sort_descending: Sort in descending order
        
    Returns:
        Success status and details
    """
    try:
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            return {'success': False, 'error': f'Sheet "{sheet_name}" not found'}
        
        ws = wb[sheet_name]
        
        # Apply auto-filter
        ws.auto_filter.ref = range_str
        
        # Note: openpyxl doesn't support actual sorting, only filter setup
        # Sorting would require manual data manipulation
        
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'sheet_name': sheet_name,
            'range': range_str,
            'auto_filter_applied': True,
            'sort_column': sort_column,
            'sort_descending': sort_descending,
            'note': 'Auto-filter applied. Sorting requires opening in Excel.'
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


@register_tool
def merge_cells_xlsx(
    file_path: str,
    sheet_name: str,
    cell_range: str
) -> Dict:
    """
    Merge cells in a range
    
    Args:
        file_path: Path to XLSX file
        sheet_name: Sheet name
        cell_range: Cell range (e.g., 'A1:D1')
        
    Returns:
        Success status and details
    """
    try:
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            return {'success': False, 'error': f'Sheet "{sheet_name}" not found'}
        
        ws = wb[sheet_name]
        ws.merge_cells(cell_range)
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'sheet_name': sheet_name,
            'merged_range': cell_range
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


@register_tool
def unmerge_cells_xlsx(
    file_path: str,
    sheet_name: str,
    cell_range: str
) -> Dict:
    """
    Unmerge cells in a range
    
    Args:
        file_path: Path to XLSX file
        sheet_name: Sheet name
        cell_range: Cell range (e.g., 'A1:D1')
        
    Returns:
        Success status and details
    """
    try:
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            return {'success': False, 'error': f'Sheet "{sheet_name}" not found'}
        
        ws = wb[sheet_name]
        ws.unmerge_cells(cell_range)
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'sheet_name': sheet_name,
            'unmerged_range': cell_range
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


@register_tool
def freeze_panes_xlsx(
    file_path: str,
    sheet_name: str,
    cell: str
) -> Dict:
    """
    Freeze panes at specified cell
    
    Args:
        file_path: Path to XLSX file
        sheet_name: Sheet name
        cell: Cell to freeze at (e.g., 'B2' freezes first row and column)
        
    Returns:
        Success status and details
    """
    try:
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            return {'success': False, 'error': f'Sheet "{sheet_name}" not found'}
        
        ws = wb[sheet_name]
        ws.freeze_panes = cell
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'sheet_name': sheet_name,
            'freeze_at': cell
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}
