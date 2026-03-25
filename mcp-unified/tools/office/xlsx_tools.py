"""
XLSX Tools for reading and writing Microsoft Excel spreadsheets
"""
from pathlib import Path
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tools.base import BaseTool, register_tool


@register_tool
def read_xlsx(file_path: str, sheet_name: Optional[str] = None) -> Dict:
    """
    Read an XLSX file and return its content
    
    Args:
        file_path: Path to the .xlsx file
        sheet_name: Specific sheet to read (None for all sheets)
        
    Returns:
        Dictionary with sheet data and metadata
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        
        sheets_data = {}
        sheets_to_read = [sheet_name] if sheet_name else wb.sheetnames
        
        for s_name in sheets_to_read:
            if s_name not in wb.sheetnames:
                continue
                
            ws = wb[s_name]
            sheet_data = {
                'rows': [],
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'headers': []
            }
            
            # Read all rows
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                sheet_data['rows'].append(row_data)
            
            # Extract headers (first row)
            if sheet_data['rows']:
                sheet_data['headers'] = sheet_data['rows'][0]
            
            sheets_data[s_name] = sheet_data
        
        return {
            'success': True,
            'file_path': file_path,
            'sheets': sheets_data,
            'sheet_names': wb.sheetnames,
            'total_sheets': len(wb.sheetnames)
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path
        }


@register_tool
def write_xlsx(
    file_path: str,
    data: Dict[str, List[List[Any]]],
    sheet_titles: Optional[Dict[str, str]] = None
) -> Dict:
    """
    Create a new XLSX file with specified data
    
    Args:
        file_path: Path where the .xlsx file will be saved
        data: Dictionary with sheet names as keys and 2D arrays as values
        sheet_titles: Optional dictionary mapping sheet names to display titles
        
    Returns:
        Success status and file path
    """
    try:
        wb = Workbook()
        first_sheet = True
        
        for sheet_name, sheet_data in data.items():
            # Create new sheet or use default
            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            # Add data to sheet
            for row_idx, row_data in enumerate(sheet_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'message': f'Excel file saved with {len(data)} sheets',
            'sheets': list(data.keys())
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path
        }


@register_tool
def extract_data_xlsx(
    file_path: str,
    sheet_name: str = None,
    has_headers: bool = True,
    row_range: Optional[tuple] = None
) -> Dict:
    """
    Extract data from XLSX file with various options
    
    Args:
        file_path: Path to the .xlsx file
        sheet_name: Sheet to read (None for first sheet)
        has_headers: Whether first row contains headers
        row_range: Optional tuple (start, end) for row range
        
    Returns:
        Dictionary with extracted data in various formats
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        
        # Get sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title
        
        # Determine row range
        start_row, end_row = row_range if row_range else (1, ws.max_row)
        
        # Read data
        raw_data = []
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            row_data = [cell if cell is not None else '' for cell in row]
            raw_data.append(row_data)
        
        # Process based on headers
        headers = raw_data[0] if has_headers and raw_data else []
        data_rows = raw_data[1:] if has_headers else raw_data
        
        # Create records (list of dicts)
        records = []
        if headers:
            for row in data_rows:
                record = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        record[header] = row[idx]
                records.append(record)
        
        return {
            'success': True,
            'file_path': file_path,
            'sheet_name': sheet_name,
            'headers': headers,
            'data': records if headers else data_rows,
            'raw_data': raw_data,
            'total_rows': len(data_rows),
            'total_columns': len(headers) if headers else (len(raw_data[0]) if raw_data else 0)
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path
        }


@register_tool
def edit_xlsx(file_path: str, edits: List[Dict]) -> Dict:
    """
    Edit an existing XLSX file
    
    Args:
        file_path: Path to the .xlsx file
        edits: List of edit operations
               Each edit should have:
               - 'sheet': target sheet name
               - 'action': 'update_cell', 'insert_row', 'delete_row', 'add_sheet', 'delete_sheet'
               - Additional parameters based on action
               
    Returns:
        Success status and details of changes
    """
    try:
        wb = load_workbook(file_path)
        changes_made = []
        
        for edit in edits:
            sheet_name = edit.get('sheet', wb.active.title)
            action = edit.get('action')
            
            if sheet_name not in wb.sheetnames:
                changes_made.append(f"Sheet '{sheet_name}' not found")
                continue
            
            ws = wb[sheet_name]
            
            if action == 'update_cell':
                row = edit.get('row', 1)
                col = edit.get('column', 1)
                value = edit.get('value', '')
                ws.cell(row=row, column=col, value=value)
                changes_made.append(f"Updated cell ({row}, {col}) in '{sheet_name}'")
                
            elif action == 'insert_row':
                row = edit.get('row', 1)
                values = edit.get('values', [])
                ws.insert_rows(row)
                for col_idx, value in enumerate(values, 1):
                    ws.cell(row=row, column=col_idx, value=value)
                changes_made.append(f"Inserted row {row} in '{sheet_name}'")
                
            elif action == 'delete_row':
                row = edit.get('row', 1)
                ws.delete_rows(row)
                changes_made.append(f"Deleted row {row} from '{sheet_name}'")
                
            elif action == 'add_sheet':
                new_sheet = edit.get('new_sheet_name', 'New Sheet')
                wb.create_sheet(title=new_sheet)
                changes_made.append(f"Added sheet '{new_sheet}'")
                
            elif action == 'delete_sheet':
                if sheet_name in wb.sheetnames and len(wb.sheetnames) > 1:
                    del wb[sheet_name]
                    changes_made.append(f"Deleted sheet '{sheet_name}'")
        
        # Save changes
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'changes_made': changes_made,
            'total_changes': len(changes_made),
            'sheets': wb.sheetnames
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path
        }


@register_tool
def format_xlsx(file_path: str, formats: List[Dict]) -> Dict:
    """
    Apply formatting to XLSX file
    
    Args:
        file_path: Path to the .xlsx file
        formats: List of formatting operations
                Each format should have:
                - 'sheet': target sheet name
                - 'range': cell range (e.g., 'A1:D10')
                - 'format_type': 'bold', 'color', 'alignment', 'border'
                - Additional parameters based on format_type
                
    Returns:
        Success status and formatting details
    """
    try:
        wb = load_workbook(file_path)
        formats_applied = []
        
        for fmt in formats:
            sheet_name = fmt.get('sheet', wb.active.title)
            
            if sheet_name not in wb.sheetnames:
                formats_applied.append(f"Sheet '{sheet_name}' not found")
                continue
            
            ws = wb[sheet_name]
            cell_range = fmt.get('range', 'A1')
            format_type = fmt.get('format_type')
            
            # Apply formatting to range
            for row in ws[cell_range]:
                for cell in row:
                    if format_type == 'bold':
                        cell.font = Font(bold=True)
                    elif format_type == 'italic':
                        cell.font = Font(italic=True)
                    elif format_type == 'color':
                        color = fmt.get('color', 'FFFF00')
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    elif format_type == 'center':
                        cell.alignment = Alignment(horizontal='center')
                    elif format_type == 'border':
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.border = thin_border
            
            formats_applied.append(f"Applied {format_type} to {cell_range} in '{sheet_name}'")
        
        # Save changes
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'formats_applied': formats_applied,
            'total_formats': len(formats_applied)
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path
        }


@register_tool
def import_csv_xlsx(
    csv_path: str,
    xlsx_path: str,
    sheet_name: Optional[str] = None,
    delimiter: str = ',',
    encoding: str = 'utf-8',
    has_headers: bool = True
) -> Dict:
    """
    Import CSV file to Excel format
    
    Args:
        csv_path: Path to the source CSV file
        xlsx_path: Path where the .xlsx file will be saved
        sheet_name: Name for the sheet (default: 'Sheet1')
        delimiter: CSV delimiter character (default: ',')
        encoding: File encoding (default: 'utf-8')
        has_headers: Whether first row contains headers (default: True)
        
    Returns:
        Success status and import details
    """
    try:
        import csv
        
        # Read CSV file
        with open(csv_path, 'r', encoding=encoding, newline='') as f:
            reader = csv.reader(f, delimiter=delimiter)
            rows = list(reader)
        
        if not rows:
            return {
                'success': False,
                'error': 'CSV file is empty',
                'csv_path': csv_path
            }
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name if sheet_name else 'Sheet1'
        
        # Add data to worksheet
        for row_data in rows:
            ws.append(row_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(xlsx_path)
        
        return {
            'success': True,
            'csv_path': csv_path,
            'xlsx_path': xlsx_path,
            'sheet_name': ws.title,
            'rows_imported': len(rows),
            'columns_imported': len(rows[0]) if rows else 0,
            'has_headers': has_headers
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'csv_path': csv_path
        }


@register_tool
def export_to_csv_xlsx(
    xlsx_path: str,
    csv_path: str,
    sheet_name: Optional[str] = None,
    delimiter: str = ',',
    encoding: str = 'utf-8',
    include_headers: bool = True
) -> Dict:
    """
    Export Excel sheet to CSV format
    
    Args:
        xlsx_path: Path to the source .xlsx file
        csv_path: Path where the .csv file will be saved
        sheet_name: Sheet to export (default: first sheet)
        delimiter: CSV delimiter character (default: ',')
        encoding: File encoding (default: 'utf-8')
        include_headers: Whether to include header row (default: True)
        
    Returns:
        Success status and export details
    """
    try:
        import csv
        
        # Load workbook
        wb = load_workbook(xlsx_path, data_only=True)
        
        # Get sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title
        
        # Read data from worksheet
        data = []
        for row in ws.iter_rows(values_only=True):
            # Convert None to empty string
            row_data = [str(cell) if cell is not None else '' for cell in row]
            data.append(row_data)
        
        # Skip headers if not needed
        if not include_headers and data:
            data = data[1:]
        
        # Write to CSV
        with open(csv_path, 'w', encoding=encoding, newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerows(data)
        
        return {
            'success': True,
            'xlsx_path': xlsx_path,
            'csv_path': csv_path,
            'sheet_name': sheet_name,
            'rows_exported': len(data),
            'columns_exported': len(data[0]) if data else 0,
            'include_headers': include_headers
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'xlsx_path': xlsx_path
        }


@register_tool
def set_cell_formula_xlsx(
    file_path: str,
    sheet_name: str,
    cell: str,
    formula: str,
    calculate: bool = False
) -> Dict:
    """
    Set formula for a cell in XLSX
    
    Args:
        file_path: Path to the .xlsx file
        sheet_name: Name of the sheet
        cell: Cell reference (e.g., 'A1', 'B2', 'C3')
        formula: Excel formula (e.g., '=SUM(A1:A10)', '=A1+B1', '=AVERAGE(B2:B20)')
        calculate: Whether to calculate the formula result (default: False)
                  Note: Formula calculation requires data_only=False when loading
        
    Returns:
        Success status and formula details
    """
    try:
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            return {
                'success': False,
                'error': f'Sheet "{sheet_name}" not found. Available sheets: {wb.sheetnames}'
            }
        
        ws = wb[sheet_name]
        
        # Validate formula starts with =
        if not formula.startswith('='):
            return {
                'success': False,
                'error': 'Formula must start with "="'
            }
        
        # Set formula
        ws[cell].value = formula
        
        # Note: Formula calculation is complex and may require external libraries
        # or specific Excel engine. For now, we just set the formula.
        calculated_value = None
        if calculate:
            # Try to get calculated value from data_only load
            try:
                wb_data = load_workbook(file_path, data_only=True)
                if sheet_name in wb_data.sheetnames:
                    calculated_value = wb_data[sheet_name][cell].value
            except:
                pass
        
        # Save workbook
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'sheet_name': sheet_name,
            'cell': cell,
            'formula': formula,
            'calculated_value': calculated_value,
            'calculate_requested': calculate
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'file_path': file_path
        }
