#!/usr/bin/env python3
"""
Document Management System - OCR Engine
=======================================
OCR processor menggunakan PaddleOCR untuk PDF dan gambar.
"""

import io
from pathlib import Path
from typing import Dict, Optional, List
from dataclasses import dataclass
import time


@dataclass
class OCRResult:
    """OCR processing result"""
    text: str
    confidence: float
    language: str
    engine: str
    processing_time: float
    page_count: int = 1


class OCREngine:
    """OCR engine using PaddleOCR"""
    
    def __init__(self, lang: str = 'id', use_gpu: bool = False):
        self.lang = lang
        self.use_gpu = use_gpu
        self.engine = None
        self._initialized = False
    
    def _init_engine(self):
        """Initialize PaddleOCR engine"""
        if self._initialized:
            return
        
        try:
            from paddleocr import PaddleOCR
            
            self.engine = PaddleOCR(
                use_angle_cls=True,
                lang=self.lang,
                use_gpu=self.use_gpu,
                show_log=False
            )
            self._initialized = True
            print(f"✅ PaddleOCR initialized (lang={self.lang}, gpu={self.use_gpu})")
            
        except ImportError:
            print("❌ PaddleOCR not installed. Install with: pip install paddleocr")
            raise
    
    def process_image(self, image_path: Path) -> Optional[OCRResult]:
        """Process single image file"""
        self._init_engine()
        
        start_time = time.time()
        
        try:
            result = self.engine.ocr(str(image_path), cls=True)
            
            texts = []
            confidences = []
            
            if result and result[0]:
                for line in result[0]:
                    if line:
                        text = line[1][0]
                        conf = line[1][1]
                        texts.append(text)
                        confidences.append(conf)
            
            full_text = '\n'.join(texts)
            avg_confidence = sum(confidences) / len(confidences) if confidences else 0.0
            processing_time = time.time() - start_time
            
            return OCRResult(
                text=full_text,
                confidence=round(avg_confidence, 3),
                language=self.lang,
                engine='paddleocr',
                processing_time=round(processing_time, 2),
                page_count=1
            )
            
        except Exception as e:
            print(f"❌ OCR error for {image_path}: {e}")
            return None
    
    def process_pdf(self, pdf_path: Path, dpi: int = 300) -> Optional[OCRResult]:
        """Process PDF file (convert to images then OCR)"""
        self._init_engine()
        
        start_time = time.time()
        
        try:
            from pdf2image import convert_from_path
            import tempfile
            
            # Convert PDF to images
            print(f"📄 Converting PDF to images: {pdf_path.name}")
            images = convert_from_path(pdf_path, dpi=dpi)
            
            all_texts = []
            all_confidences = []
            
            with tempfile.TemporaryDirectory() as temp_dir:
                for i, image in enumerate(images):
                    # Save temp image
                    temp_path = Path(temp_dir) / f"page_{i+1}.png"
                    image.save(temp_path, 'PNG')
                    
                    # OCR the page
                    result = self.engine.ocr(str(temp_path), cls=True)
                    
                    if result and result[0]:
                        page_texts = []
                        page_confidences = []
                        
                        for line in result[0]:
                            if line:
                                text = line[1][0]
                                conf = line[1][1]
                                page_texts.append(text)
                                page_confidences.append(conf)
                        
                        if page_texts:
                            all_texts.append(f"--- Page {i+1} ---")
                            all_texts.extend(page_texts)
                            all_confidences.extend(page_confidences)
            
            full_text = '\n'.join(all_texts)
            avg_confidence = sum(all_confidences) / len(all_confidences) if all_confidences else 0.0
            processing_time = time.time() - start_time
            
            return OCRResult(
                text=full_text,
                confidence=round(avg_confidence, 3),
                language=self.lang,
                engine='paddleocr',
                processing_time=round(processing_time, 2),
                page_count=len(images)
            )
            
        except ImportError:
            print("❌ pdf2image not installed. Install with: pip install pdf2image")
            print("   Also need: sudo apt-get install poppler-utils")
            return None
        except Exception as e:
            print(f"❌ PDF OCR error for {pdf_path}: {e}")
            return None
    
    def process_bytes(self, file_bytes: bytes, file_ext: str) -> Optional[OCRResult]:
        """Process file from bytes"""
        import tempfile
        
        with tempfile.NamedTemporaryFile(suffix=file_ext, delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = Path(tmp.name)
        
        try:
            if file_ext.lower() in ['.pdf']:
                result = self.process_pdf(tmp_path)
            elif file_ext.lower() in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.webp']:
                result = self.process_image(tmp_path)
            else:
                result = None
        finally:
            tmp_path.unlink(missing_ok=True)
        
        return result


class TextExtractor:
    """Extract text from various document formats"""
    
    @staticmethod
    def extract_pdf_text(pdf_path: Path) -> Optional[str]:
        """Extract text from PDF using pdfplumber (faster than OCR for text PDFs)"""
        try:
            import pdfplumber
            
            texts = []
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        texts.append(text)
            
            return '\n'.join(texts) if texts else None
            
        except ImportError:
            print("❌ pdfplumber not installed. Install with: pip install pdfplumber")
            return None
        except Exception as e:
            print(f"❌ PDF extraction error: {e}")
            return None
    
    @staticmethod
    def extract_docx_text(docx_path: Path) -> Optional[str]:
        """Extract text from DOCX"""
        try:
            from docx import Document
            
            doc = Document(docx_path)
            texts = [paragraph.text for paragraph in doc.paragraphs if paragraph.text]
            
            return '\n'.join(texts) if texts else None
            
        except ImportError:
            print("❌ python-docx not installed. Install with: pip install python-docx")
            return None
        except Exception as e:
            print(f"❌ DOCX extraction error: {e}")
            return None
    
    @staticmethod
    def extract_xlsx_text(xlsx_path: Path) -> Optional[str]:
        """Extract text from XLSX"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(xlsx_path)
            texts = []
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                texts.append(f"--- Sheet: {sheet} ---")
                
                for row in ws.iter_rows():
                    row_text = ' '.join([str(cell.value) for cell in row if cell.value])
                    if row_text.strip():
                        texts.append(row_text)
            
            return '\n'.join(texts) if texts else None
            
        except ImportError:
            print("❌ openpyxl not installed. Install with: pip install openpyxl")
            return None
        except Exception as e:
            print(f"❌ XLSX extraction error: {e}")
            return None
    
    @staticmethod
    def extract_txt_text(txt_path: Path) -> Optional[str]:
        """Extract text from TXT"""
        try:
            with open(txt_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except Exception as e:
            print(f"❌ TXT extraction error: {e}")
            return None


def process_document(file_path: Path, use_ocr: bool = True) -> Optional[OCRResult]:
    """
    Process document with smart extraction:
    1. Try native text extraction first (faster)
    2. Fall back to OCR if needed or if image-based
    """
    ext = file_path.suffix.lower()
    extractor = TextExtractor()
    
    # For images, always use OCR
    if ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.webp']:
        if use_ocr:
            ocr = OCREngine()
            return ocr.process_image(file_path)
        return None
    
    # For PDFs, try text extraction first
    if ext == '.pdf':
        text = extractor.extract_pdf_text(file_path)
        
        if text and len(text.strip()) > 100:
            # Text extraction succeeded
            return OCRResult(
                text=text,
                confidence=1.0,
                language='id',
                engine='pdfplumber',
                processing_time=0,
                page_count=text.count('--- Page') + 1
            )
        elif use_ocr:
            # Fall back to OCR
            ocr = OCREngine()
            return ocr.process_pdf(file_path)
    
    # For Office documents
    if ext == '.docx':
        text = extractor.extract_docx_text(file_path)
    elif ext in ['.xlsx', '.xls']:
        text = extractor.extract_xlsx_text(file_path)
    elif ext == '.txt':
        text = extractor.extract_txt_text(file_path)
    else:
        text = None
    
    if text:
        return OCRResult(
            text=text,
            confidence=1.0,
            language='id',
            engine='native',
            processing_time=0,
            page_count=1
        )
    
    return None


def main():
    """Test OCR engine"""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python ocr_engine.py <file_path>")
        sys.exit(1)
    
    file_path = Path(sys.argv[1])
    
    if not file_path.exists():
        print(f"❌ File not found: {file_path}")
        sys.exit(1)
    
    print(f"📝 Processing: {file_path}")
    print("=" * 60)
    
    result = process_document(file_path, use_ocr=True)
    
    if result:
        print(f"\n✅ Engine: {result.engine}")
        print(f"✅ Confidence: {result.confidence}")
        print(f"✅ Pages: {result.page_count}")
        print(f"✅ Time: {result.processing_time}s")
        print(f"\n📝 Text Preview (first 500 chars):")
        print("-" * 60)
        print(result.text[:500])
        print("..." if len(result.text) > 500 else "")
    else:
        print("❌ Failed to process document")


if __name__ == "__main__":
    main()